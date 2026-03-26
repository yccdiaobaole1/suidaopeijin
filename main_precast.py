import argparse
import csv
import json
import os
import sys
from datetime import date
import math
import hashlib
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime

# 将 src 目录加入路径
sys.path.append(os.path.join(os.path.dirname(__file__), 'src'))

from src.geometry_modeler import GeometryModeler
from src.force_mapper import ForceMapper
from src.nsga2_solver import HybridNSGA2Solver
from src.structural_solver import StructuralSolver
from src.data_loader import load_force_records, load_config, _parse_float


DEFAULT_CODE_GOVERNANCE_PARAMS: Dict[str, Any] = {
    'shear_use_abs_vm': True,
    'lambda_min': 1.5,
    'lambda_max': 3.0,
    'h_w_source': 'h0',
    'shear_n_mode': 'compression_only',
    'gamma_d_shear': 1.0,
    'rho_sv_min_hpb300': 0.0014,
    'rho_sv_min_hrb400': 0.0011,
}

DEFAULT_LOCKED_CODE_PARAMS: List[str] = [
    'shear_use_abs_vm',
    'lambda_min',
    'lambda_max',
    'h_w_source',
    'shear_n_mode',
    'gamma_d_shear',
]



def _stable_json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(',', ':'))


def _sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode('utf-8')).hexdigest()


def _sha256_file(path: str) -> str:
    try:
        with open(path, 'rb') as f:
            return hashlib.sha256(f.read()).hexdigest()
    except Exception:
        return ''


def _values_equal(left: Any, right: Any) -> bool:
    if isinstance(left, bool) or isinstance(right, bool):
        return left is right
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return abs(float(left) - float(right)) <= 1e-9
    if isinstance(left, (dict, list)) or isinstance(right, (dict, list)):
        try:
            return _stable_json_dumps(left) == _stable_json_dumps(right)
        except Exception:
            return str(left) == str(right)
    return left == right


def _format_cell_value(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, (dict, list)):
        try:
            return _stable_json_dumps(value)
        except Exception:
            return str(value)
    return str(value)


def _md_cell(value: Any) -> str:
    return _format_cell_value(value).replace('|', '\\|')


def _normalize_approval_item(item: Any) -> Optional[Dict[str, Any]]:
    if not isinstance(item, dict):
        return None
    param = str(item.get('param', '')).strip()
    if not param:
        return None
    return {
        'param': param,
        'approved': bool(item.get('approved', False)),
        'approved_by': str(item.get('approved_by', '')).strip(),
        'date': str(item.get('date', '')).strip(),
        'ticket': str(item.get('ticket', '')).strip(),
        'reason': str(item.get('reason', '')).strip(),
    }


def apply_code_governance(config: Dict[str, Any], config_file: str) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    section_input = config.get('section', {})
    if not isinstance(section_input, dict):
        raise ValueError('config.section must be an object')
    section_input = dict(section_input)

    governance_cfg = config.get('code_governance', {})
    if not isinstance(governance_cfg, dict):
        governance_cfg = {}

    project_defaults = dict(DEFAULT_CODE_GOVERNANCE_PARAMS)
    custom_defaults = governance_cfg.get('project_defaults')
    if isinstance(custom_defaults, dict):
        for key, value in custom_defaults.items():
            project_defaults[str(key)] = value

    raw_locked = governance_cfg.get('locked_params')
    if isinstance(raw_locked, list) and raw_locked:
        locked_params = [str(x).strip() for x in raw_locked if str(x).strip()]
    else:
        locked_params = list(DEFAULT_LOCKED_CODE_PARAMS)
    locked_set = set(locked_params)

    approval_map: Dict[str, List[Dict[str, Any]]] = {}
    approval_records: List[Dict[str, Any]] = []
    raw_approvals = governance_cfg.get('change_approvals', [])
    if isinstance(raw_approvals, list):
        for item in raw_approvals:
            norm = _normalize_approval_item(item)
            if not norm:
                continue
            approval_records.append(norm)
            if norm.get('approved'):
                approval_map.setdefault(norm['param'], []).append(norm)

    effective_section = dict(section_input)
    blocked_changes: List[Dict[str, Any]] = []
    approved_changes: List[Dict[str, Any]] = []
    rows: List[Dict[str, Any]] = []

    all_params = sorted(set(project_defaults.keys()) | set(section_input.keys()))
    for param in all_params:
        is_code_param = param in project_defaults
        default_value = project_defaults.get(param)
        input_exists = param in section_input
        input_value = section_input.get(param) if input_exists else None
        is_locked = is_code_param and (param in locked_set)
        approval = approval_map.get(param, [])
        latest_approval = approval[-1] if approval else None

        if is_code_param:
            if not input_exists:
                effective_value = default_value
                source = 'project_default'
            elif _values_equal(input_value, default_value):
                effective_value = input_value
                source = 'project_default'
            elif is_locked and not approval:
                effective_value = default_value
                source = 'locked_default'
                blocked_changes.append({
                    'param': param,
                    'default': default_value,
                    'requested': input_value,
                    'effective': effective_value,
                    'reason': 'locked_without_approval',
                })
            else:
                effective_value = input_value
                source = 'approved_override' if is_locked else 'config_override'
                approved_changes.append({
                    'param': param,
                    'default': default_value,
                    'effective': effective_value,
                    'source': source,
                    'approval': latest_approval,
                })
            effective_section[param] = effective_value
        else:
            effective_value = input_value
            source = 'config_input'

        rows.append({
            'param': param,
            'is_code_param': is_code_param,
            'locked': is_locked,
            'default': default_value,
            'input_exists': input_exists,
            'input': input_value,
            'effective': effective_value,
            'source': source,
            'approval': latest_approval,
        })

    effective_code_params = {k: effective_section.get(k) for k in sorted(project_defaults.keys())}
    governance_payload = {
        'mode': 'defaults+locked+approval',
        'config_file': os.path.abspath(config_file),
        'config_file_hash': _sha256_file(config_file),
        'effective_section_fingerprint': _sha256_text(_stable_json_dumps(effective_section)),
        'effective_code_fingerprint': _sha256_text(_stable_json_dumps(effective_code_params)),
        'project_defaults': project_defaults,
        'locked_params': locked_params,
        'rows': rows,
        'blocked_changes': blocked_changes,
        'approved_changes': approved_changes,
        'blocked_change_count': len(blocked_changes),
        'approved_change_count': len(approved_changes),
        'approval_records': approval_records,
    }

    governed_config = dict(config)
    governed_config['section'] = effective_section
    return governed_config, governance_payload


def _build_governance_source_summary(governance: Dict[str, Any], max_items: int = 12) -> str:
    if not governance:
        return ''
    rows = governance.get('rows', [])
    if not isinstance(rows, list):
        return ''
    changed = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        if not row.get('is_code_param'):
            continue
        source = str(row.get('source', ''))
        if source == 'project_default':
            continue
        changed.append(f"{row.get('param')}={source}")
    if not changed:
        return 'all_project_default'
    if len(changed) > max_items:
        return '; '.join(changed[:max_items]) + '; ...'
    return '; '.join(changed)


def build_segments(config: Dict[str, Any]):
    geo = config.get('geometry', {})
    clearance_w = geo.get('clearance_width', 13.3)
    clearance_h = geo.get('clearance_height', 10.88)

    modeler = GeometryModeler(clearance_w, clearance_h)
    params = geo.get('params')
    if isinstance(params, dict):
        modeler.update_parameters(**params)
    segments = modeler.partition_3A_2B_2C()
    return segments


def _format_value(value, digits=3):
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        if digits is None:
            return str(value)
        return f'{value:.{digits}f}'
    return str(value)


def _format_crack(value, digits=3):
    """格式化裂缝宽度，0.000 标注为全截面受压"""
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        if abs(value) < 1e-9:
            return f'0.{"0" * digits} (全截面受压)'
        return f'{value:.{digits}f}'
    return str(value)


def _format_stirrup_spec(stirrup: Dict[str, Any]) -> str:
    if not stirrup:
        return ''
    grade = stirrup.get('grade', '')
    dia = stirrup.get('diameter', '')
    spacing = _format_value(stirrup.get('spacing'), 1)
    legs = stirrup.get('legs', '')
    layout = stirrup.get('layout', '')
    layout_tag = f' {layout}' if layout else ''
    return f'{grade} D{dia}@{spacing} ({legs} legs{layout_tag})'


def _format_dist_spec(dist_rebar: Dict[str, Any]) -> str:
    if not dist_rebar:
        return ''
    grade = dist_rebar.get('grade', '')
    dia = dist_rebar.get('diameter', '')
    spacing = _format_value(dist_rebar.get('spacing'), 1)
    count = dist_rebar.get('count', '')
    mode = dist_rebar.get('mode', '')
    mode_tag = f' {mode}' if mode else ''
    return f'{grade} D{dia}@{spacing} ({count} bars{mode_tag})'


def _format_main_spec(main_rebar: Dict[str, Any]) -> str:
    if not main_rebar:
        return ''
    grade = main_rebar.get('grade', '')
    dia = main_rebar.get('diameter', '')
    count = main_rebar.get('count', '')
    spacing = _format_value(main_rebar.get('spacing'), 1)
    return f'{grade} D{dia} x{count} @ {spacing}'


def _to_float(value: Any, default: float = float('inf')) -> float:
    try:
        if value is None:
            return default
        return float(value)
    except Exception:
        return default


def _to_int(value: Any, default: int = 0) -> int:
    try:
        if value is None:
            return default
        return int(float(value))
    except Exception:
        return default


def _same_stirrup_spec(left: Dict[str, Any], right: Dict[str, Any]) -> bool:
    if not left or not right:
        return False
    if str(left.get('grade', '')) != str(right.get('grade', '')):
        return False
    if _to_int(left.get('diameter', 0), 0) != _to_int(right.get('diameter', 0), 0):
        return False
    if _to_int(left.get('legs', 0), 0) != _to_int(right.get('legs', 0), 0):
        return False
    if str(left.get('layout', '')) != str(right.get('layout', '')):
        return False
    return abs(_to_float(left.get('spacing'), 0.0) - _to_float(right.get('spacing'), 0.0)) <= 1e-6


def _format_delta_text(delta: Dict[str, Any]) -> str:
    if not delta:
        return ''
    return (
        f"cost={_format_value(delta.get('cost'), 2)}, "
        f"crack={_format_value(delta.get('max_crack'), 3)}, "
        f"Asv/s={_format_value(delta.get('provided_asv_over_s'), 5)}, "
        f"dist_ratio={_format_value(delta.get('dist_ratio_to_main'), 4)}, "
        f"spacing={_format_value(delta.get('main_spacing'), 1)}"
    )


def _build_joint_top_summary(joint_selection: Dict[str, Any], top_n: int = 3) -> str:
    rows = joint_selection.get('candidates', [])
    if not rows:
        return ''
    chunks = []
    for row in rows[:top_n]:
        metrics = row.get('metrics', {})
        chunks.append(
            f"#{row.get('rank')} score={_format_value(row.get('joint_score'), 4)} "
            f"cost={_format_value(metrics.get('cost'), 2)} "
            f"crack={_format_crack(metrics.get('max_crack'), 3)} "
            f"Asv/s={_format_value(metrics.get('provided_asv_over_s'), 5)} "
            f"s={_format_value(metrics.get('main_spacing'), 1)}"
        )
    return ' | '.join(chunks)


def _build_stirrup_top_summary(stirrup_selection: Dict[str, Any], top_n: int = 3) -> str:
    rows = stirrup_selection.get('candidates', [])
    if not rows:
        return ''
    chunks = []
    for row in rows[:top_n]:
        chunks.append(
            f"#{row.get('rank')} {row.get('spec', '')} "
            f"Asv/s={_format_value(row.get('provided_asv_over_s'), 5)} "
            f"dReq={_format_value(row.get('delta_to_required'), 5)}"
        )
    return ' | '.join(chunks)


def _ensure_parent_dir(path: str):
    parent = os.path.dirname(os.path.abspath(path))
    if parent and not os.path.exists(parent):
        os.makedirs(parent, exist_ok=True)


def _build_main_scheme_candidates(main_result: Dict[str, Any], max_count: int = 24) -> List[Dict[str, Any]]:
    if not main_result:
        return []

    candidates: List[Dict[str, Any]] = []
    seen = set()

    base_main = main_result.get('main_rebar', {})
    base_metrics = main_result.get('metrics', {})
    base_summary = main_result.get('summary', {})
    if base_main:
        key = (
            base_main.get('grade'),
            base_main.get('diameter'),
            int(base_main.get('count', 0)),
            round(float(base_main.get('spacing', 0.0)), 3),
        )
        seen.add(key)
        candidates.append({
            'main_rebar': dict(base_main),
            'metrics': dict(base_metrics),
            'summary': dict(base_summary),
        })

    for row in main_result.get('pareto_front', []):
        main_rebar = {
            'grade': row.get('grade'),
            'diameter': row.get('diameter'),
            'count': int(row.get('count', 0)),
            'spacing': row.get('spacing'),
        }
        key = (
            main_rebar.get('grade'),
            main_rebar.get('diameter'),
            int(main_rebar.get('count', 0)),
            round(float(main_rebar.get('spacing', 0.0)), 3),
        )
        if key in seen:
            continue
        seen.add(key)
        summary = {
            'grade': main_rebar.get('grade'),
            'diameter': main_rebar.get('diameter'),
            'count': int(main_rebar.get('count', 0)),
            'type': f"{main_rebar.get('grade')} D{main_rebar.get('diameter')}",
        }
        candidates.append({
            'main_rebar': main_rebar,
            'metrics': dict(row.get('metrics', {})),
            'summary': summary,
        })

    candidates.sort(key=lambda x: (x.get('metrics', {}).get('cost', float('inf')), x.get('metrics', {}).get('max_crack', float('inf'))))
    return candidates[:max_count]


def _joint_objectives(result: Dict[str, Any]):
    metrics = result.get('metrics', {})
    shear = result.get('shear_design', {})
    dist = result.get('dist_rebar', {})
    main = result.get('main_rebar', {})

    cost = float(metrics.get('cost', float('inf')))
    crack = float(metrics.get('max_crack', float('inf')))
    asv_over_s = float(shear.get('provided_Asv_over_s', float('inf')))
    dist_ratio = float(dist.get('ratio_to_main', float('inf')))
    spacing_obj = -float(main.get('spacing', 0.0))  # maximize spacing
    return (cost, crack, asv_over_s, dist_ratio, spacing_obj)


def _dominates(a: Dict[str, Any], b: Dict[str, Any]) -> bool:
    oa = a['joint_objs']
    ob = b['joint_objs']
    no_worse = all(x <= y for x, y in zip(oa, ob))
    strictly_better = any(x < y for x, y in zip(oa, ob))
    return no_worse and strictly_better


def _joint_pareto(candidates: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    front = []
    for i, a in enumerate(candidates):
        dominated = False
        for j, b in enumerate(candidates):
            if i == j:
                continue
            if _dominates(b, a):
                dominated = True
                break
        if not dominated:
            front.append(a)
    return front


def _normalize(value: float, low: float, high: float) -> float:
    if not math.isfinite(value):
        return 1.0
    if high <= low + 1e-12:
        return 0.0
    return (value - low) / (high - low)


def _joint_weights(preference: str = "balanced") -> List[float]:
    """联合评分权重，按偏好模式返回不同权重组合。
    5个目标: cost / crack / asv_over_s / dist_ratio / -spacing
    """
    WEIGHT_MAP = {
        "cost":         [0.55, 0.15, 0.10, 0.10, 0.10],
        "safety":       [0.15, 0.55, 0.10, 0.10, 0.10],
        "construction": [0.15, 0.10, 0.10, 0.10, 0.55],
        "balanced":     [0.45, 0.25, 0.15, 0.10, 0.05],
    }
    return WEIGHT_MAP.get(preference, WEIGHT_MAP["balanced"])


def _score_joint_front(front: List[Dict[str, Any]], preference: str = "balanced") -> List[Dict[str, Any]]:
    if not front:
        return []

    rows = []
    joint_objs = [tuple(row.get('joint_objs', (float('inf'),) * 5)) for row in front]
    cols = list(zip(*joint_objs))
    mins = [min(col) for col in cols]
    maxs = [max(col) for col in cols]
    weights = _joint_weights(preference)

    for row, objs in zip(front, joint_objs):
        main = row.get('main_rebar', {})
        stirrup = row.get('stirrup', {})
        dist = row.get('dist_rebar', {})
        metrics = row.get('metrics', {})
        shear = row.get('shear_design', {})

        cost = _to_float(metrics.get('cost'))
        crack = _to_float(metrics.get('max_crack'))
        asv_over_s = _to_float(shear.get('provided_Asv_over_s'))
        dist_ratio = _to_float(dist.get('ratio_to_main'))
        main_spacing = _to_float(main.get('spacing'), 0.0)

        normed = [_normalize(v, mins[i], maxs[i]) for i, v in enumerate(objs)]
        score = sum(weights[i] * normed[i] for i in range(len(weights)))
        rows.append({
            'raw': row,
            'joint_score': score,
            'main': {
                'grade': main.get('grade'),
                'diameter': main.get('diameter'),
                'count': main.get('count'),
                'spacing': main.get('spacing'),
            },
            'stirrup': {
                'grade': stirrup.get('grade'),
                'diameter': stirrup.get('diameter'),
                'legs': stirrup.get('legs'),
                'layout': stirrup.get('layout'),
                'spacing': stirrup.get('spacing'),
            },
            'dist': {
                'mode': dist.get('mode'),
                'grade': dist.get('grade'),
                'diameter': dist.get('diameter'),
                'count': dist.get('count'),
                'spacing': dist.get('spacing'),
            },
            'metrics': {
                'cost': cost,
                'max_crack': crack,
                'min_sf': _to_float(metrics.get('min_sf'), None),
                'main_spacing': main_spacing,
                'provided_asv_over_s': asv_over_s,
                'dist_ratio_to_main': dist_ratio,
            },
            'normed': {
                'cost': normed[0],
                'max_crack': normed[1],
                'provided_asv_over_s': normed[2],
                'dist_ratio_to_main': normed[3],
                # spacing_obj = -main_spacing，归一化值越小代表间距越大
                'main_spacing': normed[4],
            },
        })

    rows.sort(
        key=lambda x: (
            x['joint_score'],
            x['metrics']['cost'],
            x['metrics']['max_crack'],
            x['metrics']['provided_asv_over_s'],
            x['metrics']['dist_ratio_to_main'],
            -x['metrics']['main_spacing'],
        )
    )
    for idx, row in enumerate(rows, start=1):
        row['rank'] = idx
    return rows


def _select_joint_best(pareto_front: List[Dict[str, Any]], preference: str = "balanced") -> Dict[str, Any]:
    scored = _score_joint_front(pareto_front, preference)
    if not scored:
        return None
    return scored[0]['raw']


def _find_selected_joint_row(scored_rows: List[Dict[str, Any]], selected_raw: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if not scored_rows:
        return None
    if selected_raw is None:
        return None

    for row in scored_rows:
        if row.get('raw') is selected_raw:
            return row

    selected_main = selected_raw.get('main_rebar', {})
    selected_stirrup = selected_raw.get('stirrup', {})
    selected_dist = selected_raw.get('dist_rebar', {})
    for row in scored_rows:
        main = row.get('main', {})
        stirrup = row.get('stirrup', {})
        dist = row.get('dist', {})
        if (
            str(main.get('grade')) == str(selected_main.get('grade')) and
            _to_int(main.get('diameter', 0), 0) == _to_int(selected_main.get('diameter', 0), 0) and
            _to_int(main.get('count', 0), 0) == _to_int(selected_main.get('count', 0), 0) and
            abs(_to_float(main.get('spacing'), 0.0) - _to_float(selected_main.get('spacing'), 0.0)) <= 1e-6 and
            _same_stirrup_spec(stirrup, selected_stirrup) and
            str(dist.get('mode', '')) == str(selected_dist.get('mode', '')) and
            _to_int(dist.get('diameter', 0), 0) == _to_int(selected_dist.get('diameter', 0), 0) and
            _to_int(dist.get('count', 0), 0) == _to_int(selected_dist.get('count', 0), 0)
        ):
            return row
    return None


def _rank_value(values: List[float], target: float, larger_better: bool = False) -> Optional[int]:
    if not math.isfinite(target):
        return None
    finite_values = [v for v in values if math.isfinite(v)]
    if not finite_values:
        return None
    ordered = sorted(finite_values, reverse=larger_better)
    for idx, value in enumerate(ordered, start=1):
        if abs(value - target) <= 1e-9:
            return idx
    return None


def _build_joint_selection_payload(front: List[Dict[str, Any]], selected_raw: Dict[str, Any], preference: str = "balanced") -> Dict[str, Any]:
    scored = _score_joint_front(front, preference)
    weights = _joint_weights(preference)
    weight_dict = {
        'cost': weights[0],
        'max_crack': weights[1],
        'provided_asv_over_s': weights[2],
        'dist_ratio_to_main': weights[3],
        'main_spacing': weights[4],
    }
    if not scored:
        return {
            'method': 'weighted_normalized_sum',
            'preference': preference,
            'weights': weight_dict,
            'candidate_count': 0,
            'selected_rank': None,
            'selected_score': None,
            'selected_reason': 'no_joint_candidates',
            'delta_vs_runner_up': None,
            'candidates': [],
        }

    selected = _find_selected_joint_row(scored, selected_raw)
    if selected is None:
        selected = scored[0]

    candidates = []
    for row in scored:
        candidates.append({
            'rank': row.get('rank'),
            'is_selected': row is selected,
            'joint_score': row.get('joint_score'),
            'main': dict(row.get('main', {})),
            'stirrup': dict(row.get('stirrup', {})),
            'dist': dict(row.get('dist', {})),
            'metrics': dict(row.get('metrics', {})),
            'normed': dict(row.get('normed', {})),
        })

    metric_values = {
        'cost': [r['metrics']['cost'] for r in scored],
        'max_crack': [r['metrics']['max_crack'] for r in scored],
        'provided_asv_over_s': [r['metrics']['provided_asv_over_s'] for r in scored],
        'dist_ratio_to_main': [r['metrics']['dist_ratio_to_main'] for r in scored],
        'main_spacing': [r['metrics']['main_spacing'] for r in scored],
    }
    metric_ranks = {
        'cost': _rank_value(metric_values['cost'], selected['metrics']['cost'], larger_better=False),
        'max_crack': _rank_value(metric_values['max_crack'], selected['metrics']['max_crack'], larger_better=False),
        'provided_asv_over_s': _rank_value(metric_values['provided_asv_over_s'], selected['metrics']['provided_asv_over_s'], larger_better=False),
        'dist_ratio_to_main': _rank_value(metric_values['dist_ratio_to_main'], selected['metrics']['dist_ratio_to_main'], larger_better=False),
        'main_spacing': _rank_value(metric_values['main_spacing'], selected['metrics']['main_spacing'], larger_better=True),
    }

    runner_up = scored[1] if len(scored) > 1 else None
    delta_vs_runner_up = None
    if runner_up:
        delta_vs_runner_up = {
            'cost': selected['metrics']['cost'] - runner_up['metrics']['cost'],
            'max_crack': selected['metrics']['max_crack'] - runner_up['metrics']['max_crack'],
            'provided_asv_over_s': selected['metrics']['provided_asv_over_s'] - runner_up['metrics']['provided_asv_over_s'],
            'dist_ratio_to_main': selected['metrics']['dist_ratio_to_main'] - runner_up['metrics']['dist_ratio_to_main'],
            'main_spacing': selected['metrics']['main_spacing'] - runner_up['metrics']['main_spacing'],
        }

    reason = (
        f"joint_score_min={_format_value(selected.get('joint_score'), 4)}; "
        f"metric_rank(cost/crack/asv/dist/spacing)="
        f"{metric_ranks['cost']}/{metric_ranks['max_crack']}/{metric_ranks['provided_asv_over_s']}/"
        f"{metric_ranks['dist_ratio_to_main']}/{metric_ranks['main_spacing']}"
    )
    if delta_vs_runner_up:
        reason += f"; delta_vs_runner_up(selected-#2): {_format_delta_text(delta_vs_runner_up)}"

    return {
        'method': 'weighted_normalized_sum',
        'preference': preference,
        'weights': weight_dict,
        'candidate_count': len(scored),
        'selected_rank': selected.get('rank'),
        'selected_score': selected.get('joint_score'),
        'selected_reason': reason,
        'metric_ranks': metric_ranks,
        'delta_vs_runner_up': delta_vs_runner_up,
        'candidates': candidates,
    }


def _build_stirrup_selection_payload(result: Dict[str, Any]) -> Dict[str, Any]:
    stirrup = result.get('stirrup', {}) or {}
    shear = result.get('shear_design', {}) or {}
    pareto = result.get('stirrup_pareto', []) or []

    required = _to_float(shear.get('required_Asv_over_s'), 0.0)
    status = str(shear.get('status', 'unknown'))

    rows = []
    for row in pareto:
        provided = _to_float(row.get('provided_Asv_over_s'))
        dia = _to_int(row.get('diameter', 0), 0)
        spacing = _to_float(row.get('spacing'), 0.0)
        delta_req = abs(provided - required)
        spec = _format_stirrup_spec({
            'grade': row.get('grade'),
            'diameter': row.get('diameter'),
            'spacing': row.get('spacing'),
            'legs': row.get('legs'),
            'layout': row.get('layout'),
        })
        rows.append({
            'spec': spec,
            'stirrup': {
                'grade': row.get('grade'),
                'diameter': row.get('diameter'),
                'spacing': row.get('spacing'),
                'legs': row.get('legs'),
                'layout': row.get('layout'),
            },
            'provided_asv_over_s': provided,
            'max_ratio': _to_float(row.get('max_ratio')),
            'max_limit_ratio': _to_float(row.get('max_limit_ratio')),
            'delta_to_required': delta_req,
            'sort_key': (delta_req, dia, -spacing),
        })

    rows.sort(key=lambda x: x['sort_key'])
    selected = None
    for idx, row in enumerate(rows, start=1):
        row['rank'] = idx
        row['is_selected'] = _same_stirrup_spec(row.get('stirrup', {}), stirrup)
        if row['is_selected'] and selected is None:
            selected = row
    if selected is None and rows:
        selected = rows[0]
        selected['is_selected'] = True

    selected_reason = ''
    if selected:
        selected_reason = (
            f"min_delta_to_required={_format_value(selected.get('delta_to_required'), 5)}; "
            f"status={status}; "
            f"max_ratio={_format_value(selected.get('max_ratio'), 3)}; "
            f"max_limit_ratio={_format_value(selected.get('max_limit_ratio'), 3)}"
        )
    else:
        selected_reason = f'no_stirrup_pareto_candidates; status={status}'

    payload_rows = []
    for row in rows:
        payload_rows.append({
            'rank': row.get('rank'),
            'is_selected': row.get('is_selected', False),
            'spec': row.get('spec', ''),
            'provided_asv_over_s': row.get('provided_asv_over_s'),
            'delta_to_required': row.get('delta_to_required'),
            'max_ratio': row.get('max_ratio'),
            'max_limit_ratio': row.get('max_limit_ratio'),
        })

    return {
        'status': status,
        'required_asv_over_s': required,
        'provided_asv_over_s': _to_float(shear.get('provided_Asv_over_s')),
        'candidate_count': len(rows),
        'selected_rank': selected.get('rank') if selected else None,
        'selected_reason': selected_reason,
        'candidates': payload_rows,
    }


def _build_joint_pareto_payload(front: List[Dict[str, Any]], selected_raw: Dict[str, Any]) -> List[Dict[str, Any]]:
    scored = _score_joint_front(front)
    selected = _find_selected_joint_row(scored, selected_raw)
    if selected is None and scored:
        selected = scored[0]

    payload = []
    for row in scored:
        payload.append({
            'rank': row.get('rank'),
            'is_selected': row is selected,
            'joint_score': row.get('joint_score'),
            'main': dict(row.get('main', {})),
            'stirrup': dict(row.get('stirrup', {})),
            'dist': dict(row.get('dist', {})),
            'metrics': dict(row.get('metrics', {})),
            'normed': dict(row.get('normed', {})),
        })
    return payload


def write_report_md(path: str, config: Dict[str, Any], config_file: str, force_file: str, mode: str, top_k: int,
                    segments, type_source: Dict[str, str], type_results: Dict[str, Any],
                    governance: Optional[Dict[str, Any]] = None):
    _ensure_parent_dir(path)
    lines: List[str] = []
    lines.append('# 装配式明洞配筋优化报告（多偏好方案）')
    lines.append('')
    lines.append(f'- 生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    lines.append(f'- 配置文件: {os.path.basename(config_file)}')
    lines.append(f'- 内力文件: {os.path.basename(force_file)}')
    lines.append(f'- 模式: {mode}')
    if mode == 'extreme':
        lines.append(f'- top_k: {top_k}')
    lines.append('')

    # ===== SCI 新增章节：问题建模 =====
    lines.append('## 问题建模')
    lines.append('')
    lines.append('### 决策变量')
    lines.append('')
    lines.append('x = [grade ∈ {HRB400, HRB500}, d ∈ {6,...,25}mm, s ∈ {100,...,250}mm]')
    lines.append('')
    lines.append('### 目标函数')
    lines.append('')
    lines.append('```')
    lines.append('min f₁(x) = Σ(ρᵢ × Vᵢ × Pᵢ)     — 单位长度钢筋成本 (CNY/m)')
    lines.append('min f₂(x) = max(wₖ)               — 最大裂缝宽度 (mm)')
    lines.append('max f₃(x) = s                      — 主筋间距 (mm)')
    lines.append('```')
    lines.append('')

    # 检测是否有截面使用了 SF 目标
    has_sf_sections = False
    sf_section_types = []
    for stype in sorted(type_results.keys()):
        pref_data = type_results.get(stype)
        if pref_data and isinstance(pref_data, dict):
            for pref_val in pref_data.values():
                if isinstance(pref_val, dict) and pref_val.get('use_sf_objective'):
                    has_sf_sections = True
                    sf_section_types.append(stype)
                    break

    if has_sf_sections:
        lines.append('### 自适应目标策略')
        lines.append('')
        sf_types_str = '/'.join(sf_section_types)
        lines.append(f'对于全截面受压的截面类型（{sf_types_str} 型），裂缝宽度恒为零，')
        lines.append('第二目标自动从裂缝宽度切换为最小安全系数（SF），')
        lines.append('以保持三目标优化的区分度。')
        lines.append('')
        lines.append('切换判据：首代种群中所有有效解的最大裂缝宽度 < 10⁻⁹ mm。')
        lines.append('')
        lines.append(f'- 受影响截面类型: {sf_types_str}')
        lines.append(f'- 目标切换: f₂(x) = max(wₖ) → f₂(x) = -min(SF)')
        lines.append('')

    lines.append('### 约束条件')
    lines.append('')
    lines.append('```')
    lines.append('s.t.  γ₀·S ≤ R(x)                 — 承载力极限状态')
    lines.append('      wmax ≤ wlim = 0.2mm          — 裂缝宽度限值 (Q/CR 9129-2018)')
    lines.append('      ρ ≥ ρmin                      — 最小配筋率')
    lines.append('      smin ≤ s ≤ smax              — 间距构造要求')
    lines.append('      c ≥ cmin                      — 保护层厚度')
    lines.append('```')
    lines.append('')

    # ===== SCI 新增章节：算法参数 =====
    lines.append('## 算法参数')
    lines.append('')
    # 从 type_results 中提取多次运行统计信息
    sample_result = None
    sample_multi_stats = None
    for stype in sorted(type_results.keys()):
        pref_data = type_results.get(stype)
        if pref_data and isinstance(pref_data, dict):
            for pref in ["balanced", "cost", "safety", "construction"]:
                if pref in pref_data:
                    sample_result = pref_data[pref]
                    sample_multi_stats = sample_result.get('multi_run_stats')
                    break
            if sample_result:
                break
    n_runs_display = sample_multi_stats.get('n_runs', 1) if sample_multi_stats else 1
    lines.append('| 参数 | 值 | 说明 |')
    lines.append('| --- | --- | --- |')
    lines.append('| 种群规模 | 200 | NSGA-II 种群 |')
    lines.append('| 最大代数 | 200 | 进化上限 |')
    lines.append('| 交叉方式 | 均匀交叉 | 基因级别独立交叉 |')
    lines.append('| 变异方式 | 局部微扰 | ±1 索引步长 |')
    lines.append('| 收敛准则 | 5代平均成本变化率<0.1% | 提前终止 |')
    lines.append('| 搜索空间 | 2级别×10直径×16间距=320 | 可穷举 |')
    lines.append('| 并行线程 | 4 | ThreadPoolExecutor |')
    lines.append(f'| 独立运行次数 | {n_runs_display} | 统计稳健性 |')
    lines.append('| 随机种子 | 42~42+N | 可复现 |')
    lines.append('')

    # ===== SCI 新增章节：穷举对比 =====
    has_exhaustive = False
    for stype in sorted(type_results.keys()):
        pref_data = type_results.get(stype)
        if pref_data and isinstance(pref_data, dict):
            for pref_val in pref_data.values():
                if isinstance(pref_val, dict) and 'exhaustive' in pref_val:
                    has_exhaustive = True
                    break
        if has_exhaustive:
            break

    if has_exhaustive:
        lines.append('## 穷举对比')
        lines.append('')
        lines.append('| 类型 | 总组合数 | 有效解数 | 穷举Pareto规模 | NSGA-II Pareto规模 | 覆盖率 |')
        lines.append('| --- | --- | --- | --- | --- | --- |')
        for stype in sorted(type_results.keys()):
            pref_data = type_results.get(stype)
            if not pref_data or not isinstance(pref_data, dict):
                continue
            # 取任一偏好的穷举数据（所有偏好共享同一穷举结果）
            ex_info = None
            nsga_pareto_size = 0
            for pref_val in pref_data.values():
                if isinstance(pref_val, dict):
                    if 'exhaustive' in pref_val:
                        ex_info = pref_val['exhaustive']
                    jp = pref_val.get('joint_selection', {})
                    nsga_pareto_size = max(nsga_pareto_size, jp.get('candidate_count', 0))
            if ex_info:
                total = ex_info.get('total', 0)
                valid = ex_info.get('valid_count', 0)
                ex_pareto = ex_info.get('pareto_size', 0)
                coverage = f"{nsga_pareto_size}/{ex_pareto}" if ex_pareto > 0 else "-"
                lines.append(f'| {stype} | {total} | {valid} | {ex_pareto} | {nsga_pareto_size} | {coverage} |')
        lines.append('')

    # ===== SCI 新增章节：统计验证 =====
    has_multi_stats = False
    for stype in sorted(type_results.keys()):
        pref_data = type_results.get(stype)
        if pref_data and isinstance(pref_data, dict):
            for pref_val in pref_data.values():
                if isinstance(pref_val, dict) and 'multi_run_stats' in pref_val:
                    has_multi_stats = True
                    break
        if has_multi_stats:
            break

    if has_multi_stats:
        lines.append('## 统计验证')
        lines.append('')
        lines.append('| 类型 | 指标 | 均值 | 标准差 | 最优 | 最差 |')
        lines.append('| --- | --- | --- | --- | --- | --- |')
        for stype in sorted(type_results.keys()):
            pref_data = type_results.get(stype)
            if not pref_data or not isinstance(pref_data, dict):
                continue
            stats = None
            for pref_val in pref_data.values():
                if isinstance(pref_val, dict) and 'multi_run_stats' in pref_val:
                    stats = pref_val['multi_run_stats']
                    break
            if not stats:
                continue
            lines.append(f'| {stype} | 最终 HV | {_format_value(stats.get("hv_mean"), 4)} | {_format_value(stats.get("hv_std"), 4)} | {_format_value(stats.get("hv_max"), 4)} | {_format_value(stats.get("hv_min"), 4)} |')
            lines.append(f'| {stype} | 收敛代数 | {_format_value(stats.get("conv_gen_mean"), 1)} | {_format_value(stats.get("conv_gen_std"), 1)} | {_format_value(stats.get("conv_gen_min"), 0)} | {_format_value(stats.get("conv_gen_max"), 0)} |')
            lines.append(f'| {stype} | Pareto前沿规模 | {_format_value(stats.get("front_size_mean"), 1)} | {_format_value(stats.get("front_size_std"), 1)} | {_format_value(stats.get("front_size_max"), 0)} | {_format_value(stats.get("front_size_min"), 0)} |')
        lines.append('')

    lines.append('## 代表块选择')
    lines.append('')
    lines.append('| 类型 | 代表块 |')
    lines.append('| --- | --- |')
    for stype in sorted(type_source.keys()):
        lines.append(f'| {stype} | {type_source[stype]} |')
    lines.append('')

    # 偏好名称映射
    PREF_NAMES = {
        "cost": "造价优先",
        "safety": "安全优先",
        "construction": "施工优先",
        "balanced": "均衡方案",
    }
    PREFERENCES = ["cost", "safety", "construction", "balanced"]

    # 多偏好方案总览
    lines.append('## 多偏好方案总览')
    lines.append('')
    stypes_sorted = sorted(type_source.keys())

    # 判断每个截面类型的目标模式
    def _stype_uses_sf(stype):
        pref_data = type_results.get(stype)
        if pref_data and isinstance(pref_data, dict):
            for pref_val in pref_data.values():
                if isinstance(pref_val, dict):
                    return pref_val.get('use_sf_objective', False)
        return False

    header_cols = ['偏好', '权重(成本/裂缝/抗剪/分布筋/间距)']
    for stype in stypes_sorted:
        obj2_header = f'{stype}型SF' if _stype_uses_sf(stype) else f'{stype}型裂缝'
        header_cols.extend([f'{stype}型方案', f'{stype}型造价', obj2_header])
    lines.append('| ' + ' | '.join(header_cols) + ' |')
    lines.append('| ' + ' | '.join(['---'] * len(header_cols)) + ' |')
    for pref in PREFERENCES:
        weights = _joint_weights(pref)
        w_str = '/'.join(f'{w:.2f}' for w in weights)
        row_cols = [PREF_NAMES.get(pref, pref), w_str]
        for stype in stypes_sorted:
            pref_data = type_results.get(stype)
            if not pref_data or not isinstance(pref_data, dict) or pref not in pref_data:
                row_cols.extend(['无解', '-', '-'])
                continue
            result = pref_data[pref]
            summary = result.get('summary', {})
            metrics = result.get('metrics', {})
            main = result.get('main_rebar', {})
            scheme = f"{summary.get('type', '')} x{summary.get('count', '')}"
            if _stype_uses_sf(stype):
                obj2_val = _format_value(metrics.get('min_sf'), 3)
            else:
                obj2_val = _format_crack(metrics.get('max_crack'), 3)
            row_cols.extend([
                scheme,
                _format_value(metrics.get('cost'), 2),
                obj2_val,
            ])
        lines.append('| ' + ' | '.join(str(c) for c in row_cols) + ' |')
    lines.append('')

    # 按偏好分段输出
    for pref in PREFERENCES:
        pref_name = PREF_NAMES.get(pref, pref)
        weights = _joint_weights(pref)
        w_str = '/'.join(f'{w:.2f}' for w in weights)
        lines.append(f'## {pref_name} ({pref.upper()})')
        lines.append('')
        lines.append(f'- 权重(成本/裂缝/抗剪/分布筋/间距): {w_str}')
        lines.append('')

        # 类型汇总
        lines.append('### 类型汇总')
        lines.append('')
        lines.append('| 类型 | 方案 | 根数 | 主筋间距(mm) | 拉筋 | 分布筋 | 造价(CNY/m) | 裂缝/SF | 施工间距(mm) | 联合候选数 | 拉筋候选数 | 联合评分 |')
        lines.append('| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |')
        for stype in stypes_sorted:
            pref_data = type_results.get(stype)
            if not pref_data or not isinstance(pref_data, dict) or pref not in pref_data:
                lines.append(f'| {stype} | 无解 | - | - | - | - | - | - | - | - | - | - |')
                continue
            result = pref_data[pref]
            summary = result.get('summary', {})
            main = result.get('main_rebar', {})
            stirrup = result.get('stirrup', {})
            dist = result.get('dist_rebar', {})
            metrics = result.get('metrics', {})
            joint_selection = result.get('joint_selection', {})
            stirrup_selection = result.get('stirrup_selection', {})
            if _stype_uses_sf(stype):
                obj2_cell = f'SF={_format_value(metrics.get("min_sf"), 3)}'
            else:
                obj2_cell = _format_crack(metrics.get('max_crack'), 3)
            lines.append(
                f'| {stype} | {summary.get("type", "")} | '
                f'{summary.get("count", "")} | {_format_value(main.get("spacing"), 1)} | '
                f'{_format_stirrup_spec(stirrup)} | {_format_dist_spec(dist)} | '
                f'{_format_value(metrics.get("cost"), 2)} | {obj2_cell} | '
                f'{_format_value(metrics.get("spacing"), 1)} | '
                f'{joint_selection.get("candidate_count", "")} | '
                f'{stirrup_selection.get("candidate_count", "")} | '
                f'{_format_value(joint_selection.get("selected_score"), 4)} |'
            )
        lines.append('')

        # 分块结果
        lines.append('### 分块结果')
        lines.append('')
        lines.append('| 分块 | 类型 | 采用代表块 | 方案 | 根数 | 主筋间距(mm) | 拉筋 | 分布筋 | 造价(CNY/m) | 裂缝/SF |')
        lines.append('| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |')
        for seg in segments:
            seg_id = getattr(seg, 'segment_id', None)
            seg_type = getattr(seg, 'segment_type', None)
            if seg_id is None or seg_type is None:
                continue
            pref_data = type_results.get(seg_type)
            if not pref_data or not isinstance(pref_data, dict) or pref not in pref_data:
                lines.append(f'| {seg_id} | {seg_type} | {type_source.get(seg_type, "")} | 无解 | - | - | - | - | - | - |')
                continue
            result = pref_data[pref]
            summary = result.get('summary', {})
            main = result.get('main_rebar', {})
            stirrup = result.get('stirrup', {})
            dist = result.get('dist_rebar', {})
            metrics = result.get('metrics', {})
            if _stype_uses_sf(seg_type):
                obj2_cell = f'SF={_format_value(metrics.get("min_sf"), 3)}'
            else:
                obj2_cell = _format_crack(metrics.get('max_crack'), 3)
            lines.append(
                f'| {seg_id} | {seg_type} | {type_source.get(seg_type, "")} | '
                f'{summary.get("type", "")} | {summary.get("count", "")} | '
                f'{_format_value(main.get("spacing"), 1)} | {_format_stirrup_spec(stirrup)} | '
                f'{_format_dist_spec(dist)} | {_format_value(metrics.get("cost"), 2)} | '
                f'{obj2_cell} |'
            )
        lines.append('')

        # 技术复核明细（仅 balanced 输出完整候选表，其他偏好只输出选型理由）
        lines.append('### 技术复核明细')
        lines.append('')
        for stype in stypes_sorted:
            pref_data = type_results.get(stype)
            lines.append(f'#### 类型 {stype}')
            if not pref_data or not isinstance(pref_data, dict) or pref not in pref_data:
                lines.append('- 无可行解。')
                lines.append('')
                continue

            result = pref_data[pref]
            joint_selection = result.get('joint_selection', {})
            stirrup_selection = result.get('stirrup_selection', {})
            joint_delta = joint_selection.get('delta_vs_runner_up')

            lines.append(f"- 联合Pareto候选数: {joint_selection.get('candidate_count', 0)}")
            lines.append(f"- 拉筋Pareto候选数: {stirrup_selection.get('candidate_count', 0)}")
            lines.append(f"- 联合选型理由: {joint_selection.get('selected_reason', '')}")
            lines.append(f"- 拉筋选型理由: {stirrup_selection.get('selected_reason', '')}")
            lines.append(f"- 拉筋所需 Asv/s: {_format_value(stirrup_selection.get('required_asv_over_s'), 5)}")
            if joint_delta:
                lines.append(f"- 联合相对次优差值(入选-次优): {_format_delta_text(joint_delta)}")
            lines.append('')

            # 联合候选对比表
            lines.append('##### 联合候选对比')
            lines.append('')
            is_sf_type = _stype_uses_sf(stype)
            obj2_col_header = '安全系数(SF)' if is_sf_type else '最大裂缝(mm)'
            lines.append(f'| 排名 | 入选 | 主筋 | 拉筋 | 分布筋 | 造价(CNY/m) | {obj2_col_header} | Asv/s | 分布筋比 | 主筋间距(mm) | 联合评分 |')
            lines.append('| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |')
            joint_rows = joint_selection.get('candidates', [])
            if not joint_rows:
                lines.append('| - | - | - | - | - | - | - | - | - | - | - |')
            else:
                for row in joint_rows:
                    if is_sf_type:
                        obj2_val = _format_value(row.get('metrics', {}).get('min_sf'), 3)
                    else:
                        obj2_val = _format_value(row.get('metrics', {}).get('max_crack'), 3)
                    lines.append(
                        f"| {row.get('rank', '')} | {'Y' if row.get('is_selected') else ''} | "
                        f"{_format_main_spec(row.get('main', {}))} | "
                        f"{_format_stirrup_spec(row.get('stirrup', {}))} | "
                        f"{_format_dist_spec(row.get('dist', {}))} | "
                        f"{_format_value(row.get('metrics', {}).get('cost'), 2)} | "
                        f"{obj2_val} | "
                        f"{_format_value(row.get('metrics', {}).get('provided_asv_over_s'), 5)} | "
                        f"{_format_value(row.get('metrics', {}).get('dist_ratio_to_main'), 4)} | "
                        f"{_format_value(row.get('metrics', {}).get('main_spacing'), 1)} | "
                        f"{_format_value(row.get('joint_score'), 4)} |"
                    )
            lines.append('')

            # 拉筋候选对比表（所有偏好共享同一个拉筋Pareto，只在第一个偏好输出）
            if pref == PREFERENCES[0]:
                lines.append('##### 拉筋候选对比')
                lines.append('')
                lines.append('| 排名 | 入选 | 拉筋 | 提供Asv/s | 与需求差值 | max_ratio | max_limit_ratio |')
                lines.append('| --- | --- | --- | --- | --- | --- | --- |')
                stirrup_rows = stirrup_selection.get('candidates', [])
                if not stirrup_rows:
                    lines.append('| - | - | - | - | - | - | - |')
                else:
                    for row in stirrup_rows:
                        lines.append(
                            f"| {row.get('rank', '')} | {'Y' if row.get('is_selected') else ''} | "
                            f"{row.get('spec', '')} | "
                            f"{_format_value(row.get('provided_asv_over_s'), 5)} | "
                            f"{_format_value(row.get('delta_to_required'), 5)} | "
                            f"{_format_value(row.get('max_ratio'), 3)} | "
                            f"{_format_value(row.get('max_limit_ratio'), 3)} |"
                        )
                lines.append('')

    lines.append('')
    lines.append('## 规范参数治理')
    lines.append('')
    if governance:
        blocked_changes = governance.get('blocked_changes', []) or []
        approved_changes = governance.get('approved_changes', []) or []
        blocked_params = ', '.join(str(x.get('param', '')) for x in blocked_changes if isinstance(x, dict) and x.get('param'))
        if not blocked_params:
            blocked_params = 'none'
        lines.append(f"- 治理模式: {governance.get('mode', '')}")
        lines.append(f"- 配置文件哈希(SHA256): {governance.get('config_file_hash', '')}")
        lines.append(f"- 生效条文口径指纹(SHA256): {governance.get('effective_code_fingerprint', '')}")
        lines.append(f"- 生效截面参数指纹(SHA256): {governance.get('effective_section_fingerprint', '')}")
        lines.append(f"- 锁定参数数量: {len(governance.get('locked_params', []) or [])}")
        lines.append(f"- 已审批变更数量: {len(approved_changes)}")
        lines.append(f"- 被拦截变更数量: {len(blocked_changes)}")
        lines.append(f"- 被拦截参数: {blocked_params}")
        lines.append('')
        lines.append('| 参数 | 代码参数 | 锁定 | 默认值 | 输入值 | 生效值 | 来源 | 审批 |')
        lines.append('| --- | --- | --- | --- | --- | --- | --- | --- |')
        for row in governance.get('rows', []):
            if not isinstance(row, dict):
                continue
            approval = row.get('approval') if isinstance(row.get('approval'), dict) else None
            approval_text = ''
            if approval:
                approval_text = (
                    f"{approval.get('approved_by', '')}/"
                    f"{approval.get('date', '')}/"
                    f"{approval.get('ticket', '')}"
                ).strip('/')
            elif row.get('source') == 'locked_default':
                approval_text = 'blocked(no approved record)'

            input_value = row.get('input') if row.get('input_exists') else ''
            lines.append(
                f"| {_md_cell(row.get('param'))} | "
                f"{'Y' if row.get('is_code_param') else ''} | "
                f"{'Y' if row.get('locked') else ''} | "
                f"{_md_cell(row.get('default'))} | "
                f"{_md_cell(input_value)} | "
                f"{_md_cell(row.get('effective'))} | "
                f"{_md_cell(row.get('source'))} | "
                f"{_md_cell(approval_text)} |"
            )
    else:
        lines.append('- 未启用治理摘要。')
    lines.append('')

    lines.append('')
    lines.append('## 输入摘要')
    lines.append('')
    lines.append('### 截面参数')
    lines.append('')
    section = config.get('section', {})
    lines.append('| 参数 | 值 |')
    lines.append('| --- | --- |')
    for key in sorted(section.keys()):
        lines.append(f'| {_md_cell(key)} | {_md_cell(section.get(key))} |')
    lines.append('')
    lines.append('### 几何参数')
    lines.append('')
    geo = config.get('geometry', {})
    lines.append('| 参数 | 值 |')
    lines.append('| --- | --- |')
    for key in ('clearance_width', 'clearance_height'):
        if key in geo:
            lines.append(f'| {_md_cell(key)} | {_md_cell(geo.get(key))} |')
    params = geo.get('params', {})
    for key in sorted(params.keys()):
        lines.append(f'| {_md_cell(key)} | {_md_cell(params.get(key))} |')
    lines.append('')

    with open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))


def _calc_rebar_area(diameter) -> float:
    """计算单根钢筋截面积 (mm^2)"""
    d = float(diameter)
    return math.pi * d * d / 4.0


def write_paper_report_md(path: str, config: Dict[str, Any],
                          segments, type_source: Dict[str, str],
                          type_results: Dict[str, Any]):
    """生成面向论文的精简配筋优化报告（无算法参数、无技术复核明细）"""
    _ensure_parent_dir(path)
    lines: List[str] = []

    section = config.get('section', {})
    b = float(section.get('b', 1000))
    h = float(section.get('h', 700))
    cover = section.get('cover', 50)
    concrete_grade = section.get('concrete_grade', 'C50')

    PREF_NAMES = {
        "cost": "造价优先",
        "safety": "安全优先",
        "construction": "施工优先",
        "balanced": "均衡方案",
    }
    PREFERENCES = ["cost", "safety", "construction", "balanced"]
    stypes_sorted = sorted(type_source.keys())

    def _stype_uses_sf(stype):
        pref_data = type_results.get(stype)
        if pref_data and isinstance(pref_data, dict):
            for pref_val in pref_data.values():
                if isinstance(pref_val, dict):
                    return pref_val.get('use_sf_objective', False)
        return False

    # === 标题 ===
    lines.append('# 装配式隧道衬砌配筋优化结果')
    lines.append('')
    lines.append(f'- 生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    lines.append('')

    # === 第一部分：设计条件 ===
    lines.append('## 1 设计条件')
    lines.append('')
    lines.append(f'- 截面尺寸: b x h = {b:.0f} x {h:.0f} mm')
    lines.append(f'- 保护层厚度: {cover} mm')
    lines.append(f'- 混凝土等级: {concrete_grade}')
    lines.append('- 优化方法: NSGA-II 多目标优化')
    lines.append('- 搜索空间: 2 级别 x 10 直径 x 16 间距 = 320 组合')
    lines.append('')

    # === 第二部分：各偏好方案配筋详情 ===
    lines.append('## 2 各偏好方案配筋详情')
    lines.append('')

    for pref in PREFERENCES:
        pref_name = PREF_NAMES.get(pref, pref)
        lines.append(f'### 2.{PREFERENCES.index(pref) + 1} {pref_name}')
        lines.append('')

        lines.append('| 分块 | 类型 | 主筋 | 双侧根数 | 间距(mm) | As,total(mm2) | rho_total(%) '
                     '| 拉筋 | Asv/s(mm2/mm) | rho_sv(%) '
                     '| 分布筋 | As,d(mm2) | rho_d(%) '
                     '| 造价(CNY/m) | 裂缝/SF |')
        lines.append('| --- | --- | --- | --- | --- | --- | --- '
                     '| --- | --- | --- '
                     '| --- | --- | --- '
                     '| --- | --- |')

        for seg in segments:
            seg_id = getattr(seg, 'segment_id', None)
            seg_type = getattr(seg, 'segment_type', None)
            if seg_id is None or seg_type is None:
                continue
            pref_data = type_results.get(seg_type)
            if not pref_data or not isinstance(pref_data, dict) or pref not in pref_data:
                lines.append(f'| {seg_id} | {seg_type} | - | - | - | - | - | - | - | - | - | - | - | - | - |')
                continue

            result = pref_data[pref]
            main = result.get('main_rebar', {})
            stirrup = result.get('stirrup', {})
            dist = result.get('dist_rebar', {})
            metrics = result.get('metrics', {})
            shear = result.get('shear_design', {})

            # 主筋面积和配筋率（对称双层配筋，As_total = 2 × 单侧）
            main_dia = float(main.get('diameter', 0))
            main_count = int(main.get('count', 0))
            as_single = _calc_rebar_area(main_dia) * main_count
            as_total = 2.0 * as_single
            count_total = 2 * main_count
            rho_total = as_total / (b * h) * 100.0 if (b * h) > 0 else 0.0

            main_spec = f"{main.get('grade', '')} D{main.get('diameter', '')}"
            main_spacing = _format_value(main.get('spacing'), 1)

            # 拉筋 Asv/s 和配筋率
            provided_asv_s = shear.get('provided_Asv_over_s', 0.0)
            if provided_asv_s is None:
                provided_asv_s = 0.0
            stirrup_dia = float(stirrup.get('diameter', 0))
            stirrup_legs = int(stirrup.get('legs', 2))
            stirrup_spacing = float(stirrup.get('spacing', 100))
            asv_single = _calc_rebar_area(stirrup_dia) * stirrup_legs
            rho_sv = asv_single / (b * stirrup_spacing) * 100.0 if (b * stirrup_spacing) > 0 else 0.0

            stirrup_spec = _format_stirrup_spec(stirrup)

            # 分布筋面积和配筋率
            as_dist = float(dist.get('area_provided', 0) or 0)
            rho_dist = as_dist / (b * h) * 100.0 if (b * h) > 0 else 0.0
            dist_spec = _format_dist_spec(dist)

            # 造价
            cost_val = _format_value(metrics.get('cost'), 2)

            # 裂缝/SF
            if _stype_uses_sf(seg_type):
                obj2_val = f'SF={_format_value(metrics.get("min_sf"), 3)}'
            else:
                obj2_val = _format_crack(metrics.get('max_crack'), 3)

            lines.append(
                f'| {seg_id} | {seg_type} '
                f'| {main_spec} | {count_total} | {main_spacing} '
                f'| {as_total:.1f} | {rho_total:.3f} '
                f'| {stirrup_spec} | {_format_value(provided_asv_s, 4)} | {rho_sv:.4f} '
                f'| {dist_spec} | {as_dist:.1f} | {rho_dist:.3f} '
                f'| {cost_val} | {obj2_val} |'
            )
        lines.append('')

    # === 第三部分：多偏好方案对比总览 ===
    lines.append('## 3 多偏好方案对比总览')
    lines.append('')

    header_cols = ['偏好']
    for stype in stypes_sorted:
        obj2_header = f'{stype}型SF' if _stype_uses_sf(stype) else f'{stype}型裂缝'
        header_cols.extend([f'{stype}型方案', f'{stype}型造价', obj2_header])
    lines.append('| ' + ' | '.join(header_cols) + ' |')
    lines.append('| ' + ' | '.join(['---'] * len(header_cols)) + ' |')

    for pref in PREFERENCES:
        row_cols = [PREF_NAMES.get(pref, pref)]
        for stype in stypes_sorted:
            pref_data = type_results.get(stype)
            if not pref_data or not isinstance(pref_data, dict) or pref not in pref_data:
                row_cols.extend(['无解', '-', '-'])
                continue
            result = pref_data[pref]
            summary = result.get('summary', {})
            metrics = result.get('metrics', {})
            scheme = f"{summary.get('type', '')} x{summary.get('count', '')}"
            if _stype_uses_sf(stype):
                obj2_val = _format_value(metrics.get('min_sf'), 3)
            else:
                obj2_val = _format_crack(metrics.get('max_crack'), 3)
            row_cols.extend([
                scheme,
                _format_value(metrics.get('cost'), 2),
                obj2_val,
            ])
        lines.append('| ' + ' | '.join(str(c) for c in row_cols) + ' |')
    lines.append('')

    # === 第四部分：穷举对比验证 ===
    has_exhaustive = False
    for stype in stypes_sorted:
        pref_data = type_results.get(stype)
        if pref_data and isinstance(pref_data, dict):
            for pref_val in pref_data.values():
                if isinstance(pref_val, dict) and 'exhaustive' in pref_val:
                    has_exhaustive = True
                    break
        if has_exhaustive:
            break

    if has_exhaustive:
        lines.append('## 4 穷举对比验证')
        lines.append('')
        lines.append('| 类型 | 总组合数 | 有效解数 | 穷举Pareto规模 | NSGA-II Pareto规模 | 覆盖率 |')
        lines.append('| --- | --- | --- | --- | --- | --- |')
        for stype in stypes_sorted:
            pref_data = type_results.get(stype)
            if not pref_data or not isinstance(pref_data, dict):
                continue
            ex_info = None
            nsga_pareto_size = 0
            for pref_val in pref_data.values():
                if isinstance(pref_val, dict):
                    if 'exhaustive' in pref_val:
                        ex_info = pref_val['exhaustive']
                    jp = pref_val.get('joint_selection', {})
                    nsga_pareto_size = max(nsga_pareto_size, jp.get('candidate_count', 0))
            if ex_info:
                total = ex_info.get('total', 0)
                valid = ex_info.get('valid_count', 0)
                ex_pareto = ex_info.get('pareto_size', 0)
                coverage = f"{nsga_pareto_size}/{ex_pareto}" if ex_pareto > 0 else "-"
                lines.append(f'| {stype} | {total} | {valid} | {ex_pareto} | {nsga_pareto_size} | {coverage} |')
        lines.append('')

    with open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))


def write_report_csv(path: str, segments, type_source: Dict[str, str], type_results: Dict[str, Any],
                     governance: Optional[Dict[str, Any]] = None):
    _ensure_parent_dir(path)
    governance_code_fingerprint = governance.get('effective_code_fingerprint', '') if governance else ''
    governance_config_hash = governance.get('config_file_hash', '') if governance else ''
    governance_source_summary = _build_governance_source_summary(governance or {})
    governance_blocked = governance.get('blocked_change_count', '') if governance else ''
    governance_approved = governance.get('approved_change_count', '') if governance else ''

    header = [
        'preference', 'segment_id', 'segment_type', 'representative_segment',
        'scheme', 'count', 'main_spacing_mm', 'stirrup_spec', 'dist_spec',
        'cost_cny_per_m', 'max_crack_mm', 'min_sf', 'objective_mode', 'spacing_objective_mm',
        'joint_pareto_count', 'stirrup_pareto_count',
        'joint_selected_rank', 'joint_selected_score', 'joint_selected_reason', 'joint_top3',
        'stirrup_selected_rank', 'stirrup_required_asv_over_s', 'stirrup_selected_reason', 'stirrup_top3',
        'governance_code_fingerprint', 'governance_config_sha256', 'governance_source_summary',
        'governance_blocked_count', 'governance_approved_count'
    ]

    PREFERENCES = ["cost", "safety", "construction", "balanced"]

    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        empty_tail = [''] * (len(header) - 4)  # 4 = preference + seg_id + seg_type + representative
        for pref in PREFERENCES:
            for seg in segments:
                seg_id = getattr(seg, 'segment_id', None)
                seg_type = getattr(seg, 'segment_type', None)
                if seg_id is None or seg_type is None:
                    continue
                pref_data = type_results.get(seg_type)
                if not pref_data or not isinstance(pref_data, dict) or pref not in pref_data:
                    writer.writerow([pref, seg_id, seg_type, type_source.get(seg_type, '')] + empty_tail)
                    continue
                result = pref_data[pref]
                summary = result.get('summary', {})
                main = result.get('main_rebar', {})
                stirrup = result.get('stirrup', {})
                dist = result.get('dist_rebar', {})
                metrics = result.get('metrics', {})
                joint_selection = result.get('joint_selection', {})
                stirrup_selection = result.get('stirrup_selection', {})
                writer.writerow([
                    pref,
                    seg_id,
                    seg_type,
                    type_source.get(seg_type, ''),
                    summary.get('type', ''),
                    summary.get('count', ''),
                    _format_value(main.get('spacing'), 1),
                    _format_stirrup_spec(stirrup),
                    _format_dist_spec(dist),
                    _format_value(metrics.get('cost'), 2),
                    _format_crack(metrics.get('max_crack'), 3),
                    _format_value(metrics.get('min_sf'), 3),
                    result.get('objective_mode', 'crack'),
                    _format_value(metrics.get('spacing'), 1),
                    joint_selection.get('candidate_count', ''),
                    stirrup_selection.get('candidate_count', ''),
                    joint_selection.get('selected_rank', ''),
                    _format_value(joint_selection.get('selected_score'), 4),
                    joint_selection.get('selected_reason', ''),
                    _build_joint_top_summary(joint_selection, top_n=3),
                    stirrup_selection.get('selected_rank', ''),
                    _format_value(stirrup_selection.get('required_asv_over_s'), 5),
                    stirrup_selection.get('selected_reason', ''),
                    _build_stirrup_top_summary(stirrup_selection, top_n=3),
                    governance_code_fingerprint,
                    governance_config_hash,
                    governance_source_summary,
                    governance_blocked,
                    governance_approved,
                ])


def write_report_xlsx(path: str, segments, type_source: Dict[str, str], type_results: Dict[str, Any],
                      governance: Optional[Dict[str, Any]] = None):
    _ensure_parent_dir(path)
    try:
        from openpyxl import Workbook
    except Exception:
        print('openpyxl not available, skip xlsx export.')
        return

    wb = Workbook()
    governance_code_fingerprint = governance.get('effective_code_fingerprint', '') if governance else ''
    governance_config_hash = governance.get('config_file_hash', '') if governance else ''
    governance_source_summary = _build_governance_source_summary(governance or {})
    governance_blocked = governance.get('blocked_change_count', '') if governance else ''
    governance_approved = governance.get('approved_change_count', '') if governance else ''

    PREFERENCES = ["cost", "safety", "construction", "balanced"]

    # Sheet 1: 分块结果（多偏好）
    ws_segments = wb.active
    ws_segments.title = 'Segments'
    segments_header = [
        'preference', 'segment_id', 'segment_type', 'representative_segment',
        'scheme', 'count', 'main_spacing_mm', 'stirrup_spec', 'dist_spec',
        'cost_cny_per_m', 'max_crack_mm', 'min_sf', 'objective_mode', 'spacing_objective_mm',
        'joint_pareto_count', 'stirrup_pareto_count',
        'joint_selected_rank', 'joint_selected_score', 'joint_selected_reason', 'joint_top3',
        'stirrup_selected_rank', 'stirrup_required_asv_over_s', 'stirrup_selected_reason', 'stirrup_top3',
        'governance_code_fingerprint', 'governance_config_sha256', 'governance_source_summary',
        'governance_blocked_count', 'governance_approved_count'
    ]
    ws_segments.append(segments_header)
    segments_empty_tail = [''] * (len(segments_header) - 4)
    for pref in PREFERENCES:
        for seg in segments:
            seg_id = getattr(seg, 'segment_id', None)
            seg_type = getattr(seg, 'segment_type', None)
            if seg_id is None or seg_type is None:
                continue
            pref_data = type_results.get(seg_type)
            if not pref_data or not isinstance(pref_data, dict) or pref not in pref_data:
                ws_segments.append([pref, seg_id, seg_type, type_source.get(seg_type, '')] + segments_empty_tail)
                continue
            result = pref_data[pref]
            summary = result.get('summary', {})
            main = result.get('main_rebar', {})
            stirrup = result.get('stirrup', {})
            dist = result.get('dist_rebar', {})
            metrics = result.get('metrics', {})
            joint_selection = result.get('joint_selection', {})
            stirrup_selection = result.get('stirrup_selection', {})
            ws_segments.append([
                pref,
                seg_id,
                seg_type,
                type_source.get(seg_type, ''),
                summary.get('type', ''),
                summary.get('count', ''),
                _format_value(main.get('spacing'), 1),
                _format_stirrup_spec(stirrup),
                _format_dist_spec(dist),
                _format_value(metrics.get('cost'), 2),
                _format_crack(metrics.get('max_crack'), 3),
                _format_value(metrics.get('min_sf'), 3),
                result.get('objective_mode', 'crack'),
                _format_value(metrics.get('spacing'), 1),
                joint_selection.get('candidate_count', ''),
                stirrup_selection.get('candidate_count', ''),
                joint_selection.get('selected_rank', ''),
                _format_value(joint_selection.get('selected_score'), 4),
                joint_selection.get('selected_reason', ''),
                _build_joint_top_summary(joint_selection, top_n=3),
                stirrup_selection.get('selected_rank', ''),
                _format_value(stirrup_selection.get('required_asv_over_s'), 5),
                stirrup_selection.get('selected_reason', ''),
                _build_stirrup_top_summary(stirrup_selection, top_n=3),
                governance_code_fingerprint,
                governance_config_hash,
                governance_source_summary,
                governance_blocked,
                governance_approved,
            ])

    # Sheet 2: 类型汇总（多偏好）
    ws_types = wb.create_sheet(title='Types')
    types_header = [
        'preference', 'type', 'scheme', 'count', 'main_spacing_mm', 'stirrup_spec', 'dist_spec',
        'cost_cny_per_m', 'max_crack_mm', 'min_sf', 'objective_mode', 'spacing_objective_mm',
        'joint_pareto_count', 'stirrup_pareto_count',
        'joint_selected_rank', 'joint_selected_score', 'joint_selected_reason', 'joint_top3',
        'stirrup_selected_rank', 'stirrup_required_asv_over_s', 'stirrup_selected_reason', 'stirrup_top3',
        'governance_code_fingerprint', 'governance_config_sha256', 'governance_source_summary',
        'governance_blocked_count', 'governance_approved_count'
    ]
    ws_types.append(types_header)
    types_empty_tail = [''] * (len(types_header) - 3)
    for pref in PREFERENCES:
        for stype in sorted(type_results.keys()):
            pref_data = type_results.get(stype)
            if not pref_data or not isinstance(pref_data, dict) or pref not in pref_data:
                ws_types.append([pref, stype, '无解'] + types_empty_tail)
                continue
            result = pref_data[pref]
            summary = result.get('summary', {})
            main = result.get('main_rebar', {})
            stirrup = result.get('stirrup', {})
            dist = result.get('dist_rebar', {})
            metrics = result.get('metrics', {})
            joint_selection = result.get('joint_selection', {})
            stirrup_selection = result.get('stirrup_selection', {})
            ws_types.append([
                pref,
                stype,
                summary.get('type', ''),
                summary.get('count', ''),
                _format_value(main.get('spacing'), 1),
                _format_stirrup_spec(stirrup),
                _format_dist_spec(dist),
                _format_value(metrics.get('cost'), 2),
                _format_crack(metrics.get('max_crack'), 3),
                _format_value(metrics.get('min_sf'), 3),
                result.get('objective_mode', 'crack'),
                _format_value(metrics.get('spacing'), 1),
                joint_selection.get('candidate_count', ''),
                stirrup_selection.get('candidate_count', ''),
                joint_selection.get('selected_rank', ''),
                _format_value(joint_selection.get('selected_score'), 4),
                joint_selection.get('selected_reason', ''),
                _build_joint_top_summary(joint_selection, top_n=3),
                stirrup_selection.get('selected_rank', ''),
                _format_value(stirrup_selection.get('required_asv_over_s'), 5),
                stirrup_selection.get('selected_reason', ''),
                _build_stirrup_top_summary(stirrup_selection, top_n=3),
                governance_code_fingerprint,
                governance_config_hash,
                governance_source_summary,
                governance_blocked,
                governance_approved,
            ])

    # Sheet 3: 代表块
    ws_rep = wb.create_sheet(title='Representative')
    ws_rep.append(['type', 'representative_segment'])
    for stype in sorted(type_source.keys()):
        ws_rep.append([stype, type_source[stype]])

    # Sheet 4: 联合 Pareto 细项（多偏好）
    ws_joint = wb.create_sheet(title='JointPareto')
    ws_joint.append([
        'preference', 'type', 'rank', 'is_selected',
        'main_spec', 'stirrup_spec', 'dist_spec',
        'cost_cny_per_m', 'max_crack_mm', 'min_sf', 'provided_asv_over_s', 'dist_ratio_to_main', 'main_spacing_mm',
        'joint_score', 'norm_cost', 'norm_crack', 'norm_asv_over_s', 'norm_dist_ratio', 'norm_spacing_obj'
    ])
    for pref in PREFERENCES:
        for stype in sorted(type_results.keys()):
            pref_data = type_results.get(stype)
            if not pref_data or not isinstance(pref_data, dict) or pref not in pref_data:
                continue
            result = pref_data[pref]
            joint_selection = result.get('joint_selection', {})
            for row in joint_selection.get('candidates', []):
                ws_joint.append([
                    pref,
                    stype,
                    row.get('rank', ''),
                    'Y' if row.get('is_selected') else '',
                    _format_main_spec(row.get('main', {})),
                    _format_stirrup_spec(row.get('stirrup', {})),
                    _format_dist_spec(row.get('dist', {})),
                    _format_value(row.get('metrics', {}).get('cost'), 2),
                    _format_value(row.get('metrics', {}).get('max_crack'), 3),
                    _format_value(row.get('metrics', {}).get('min_sf'), 3),
                    _format_value(row.get('metrics', {}).get('provided_asv_over_s'), 5),
                    _format_value(row.get('metrics', {}).get('dist_ratio_to_main'), 4),
                    _format_value(row.get('metrics', {}).get('main_spacing'), 1),
                    _format_value(row.get('joint_score'), 4),
                    _format_value(row.get('normed', {}).get('cost'), 4),
                    _format_value(row.get('normed', {}).get('max_crack'), 4),
                    _format_value(row.get('normed', {}).get('provided_asv_over_s'), 4),
                    _format_value(row.get('normed', {}).get('dist_ratio_to_main'), 4),
                    _format_value(row.get('normed', {}).get('main_spacing'), 4),
                ])

    # Sheet 5: 拉筋 Pareto 细项（按类型，不按偏好——拉筋候选与偏好无关）
    ws_stirrup = wb.create_sheet(title='StirrupPareto')
    ws_stirrup.append([
        'type', 'rank', 'is_selected', 'stirrup_spec',
        'provided_asv_over_s', 'delta_to_required', 'max_ratio', 'max_limit_ratio'
    ])
    for stype in sorted(type_results.keys()):
        pref_data = type_results.get(stype)
        if not pref_data or not isinstance(pref_data, dict):
            continue
        # 取 balanced 偏好的拉筋候选（所有偏好共享同一拉筋 Pareto）
        result = pref_data.get('balanced') or next(iter(pref_data.values()), None)
        if not result:
            continue
        stirrup_selection = result.get('stirrup_selection', {})
        for row in stirrup_selection.get('candidates', []):
            ws_stirrup.append([
                stype,
                row.get('rank', ''),
                'Y' if row.get('is_selected') else '',
                row.get('spec', ''),
                _format_value(row.get('provided_asv_over_s'), 5),
                _format_value(row.get('delta_to_required'), 5),
                _format_value(row.get('max_ratio'), 3),
                _format_value(row.get('max_limit_ratio'), 3),
            ])

    # Sheet 6: 选型理由（多偏好）
    ws_reason = wb.create_sheet(title='Rationale')
    ws_reason.append([
        'preference', 'type', 'joint_reason', 'stirrup_reason',
        'joint_delta_cost', 'joint_delta_crack', 'joint_delta_asv_over_s', 'joint_delta_dist_ratio', 'joint_delta_spacing_mm'
    ])
    for pref in PREFERENCES:
        for stype in sorted(type_results.keys()):
            pref_data = type_results.get(stype)
            if not pref_data or not isinstance(pref_data, dict) or pref not in pref_data:
                continue
            result = pref_data[pref]
            joint_selection = result.get('joint_selection', {})
            stirrup_selection = result.get('stirrup_selection', {})
            delta = joint_selection.get('delta_vs_runner_up') or {}
            ws_reason.append([
                pref,
                stype,
                joint_selection.get('selected_reason', ''),
                stirrup_selection.get('selected_reason', ''),
                _format_value(delta.get('cost'), 2),
                _format_value(delta.get('max_crack'), 3),
                _format_value(delta.get('provided_asv_over_s'), 5),
                _format_value(delta.get('dist_ratio_to_main'), 4),
                _format_value(delta.get('main_spacing'), 1),
            ])

    # Sheet 7: 参数治理
    ws_governance = wb.create_sheet(title='Governance')
    ws_governance.append(['item', 'value'])
    ws_governance.append(['mode', governance.get('mode', '') if governance else ''])
    ws_governance.append(['config_file', governance.get('config_file', '') if governance else ''])
    ws_governance.append(['config_file_hash', governance_config_hash])
    ws_governance.append(['effective_code_fingerprint', governance_code_fingerprint])
    ws_governance.append(['effective_section_fingerprint', governance.get('effective_section_fingerprint', '') if governance else ''])
    ws_governance.append(['locked_param_count', len(governance.get('locked_params', []) or []) if governance else 0])
    ws_governance.append(['approved_change_count', governance_approved])
    ws_governance.append(['blocked_change_count', governance_blocked])
    ws_governance.append(['source_summary', governance_source_summary])
    ws_governance.append([])
    ws_governance.append([
        'param', 'is_code_param', 'locked', 'source',
        'default', 'input', 'effective',
        'approval_by', 'approval_date', 'approval_ticket', 'approval_reason'
    ])
    if governance and isinstance(governance.get('rows'), list):
        for row in governance.get('rows', []):
            if not isinstance(row, dict):
                continue
            approval = row.get('approval') if isinstance(row.get('approval'), dict) else {}
            input_value = row.get('input') if row.get('input_exists') else ''
            ws_governance.append([
                _format_cell_value(row.get('param')),
                'Y' if row.get('is_code_param') else '',
                'Y' if row.get('locked') else '',
                _format_cell_value(row.get('source')),
                _format_cell_value(row.get('default')),
                _format_cell_value(input_value),
                _format_cell_value(row.get('effective')),
                _format_cell_value(approval.get('approved_by', '')),
                _format_cell_value(approval.get('date', '')),
                _format_cell_value(approval.get('ticket', '')),
                _format_cell_value(approval.get('reason', '')),
            ])

    wb.save(path)


def main():
    today = date.today().strftime('%Y-%m-%d')
    default_report_dir = os.path.join('reports', today)

    parser = argparse.ArgumentParser(description='Precast tunnel reinforcement optimizer')
    parser.add_argument('--config', required=True, help='path to precast config json')
    parser.add_argument('--force-file', required=True, help='path to force records (json/csv)')
    parser.add_argument('--mode', default='all', choices=['all', 'extreme'], help='force envelope reduction mode')
    parser.add_argument('--top-k', type=int, default=0, help='top-k abs(M) points when mode=extreme')
    parser.add_argument('--safety-level', type=int, default=2, choices=[1, 2], help='structural safety level')
    parser.add_argument('--report-md', default=os.path.join(default_report_dir, 'report.md'), help='markdown report output path')
    parser.add_argument('--report-csv', default=os.path.join(default_report_dir, 'report.csv'), help='csv report output path')
    parser.add_argument('--report-xlsx', default='', help='xlsx report output path (requires openpyxl)')
    parser.add_argument('--exhaustive', action='store_true', default=True, help='run exhaustive enumeration for ground truth')
    parser.add_argument('--no-exhaustive', action='store_false', dest='exhaustive', help='skip exhaustive enumeration')
    parser.add_argument('--n-runs', type=int, default=1, help='number of independent NSGA-II runs (SCI: 10)')
    parser.add_argument('--seed', type=int, default=42, help='random seed base')
    args = parser.parse_args()

    # 自动创建报告目录
    for p in [args.report_md, args.report_csv] + ([args.report_xlsx] if args.report_xlsx else []):
        d = os.path.dirname(p)
        if d:
            os.makedirs(d, exist_ok=True)

    config = load_config(args.config)
    config, governance = apply_code_governance(config, args.config)
    section_params = config.get('section', {})
    if not section_params:
        raise ValueError('config missing section params')

    if governance.get('blocked_change_count', 0):
        blocked_params = [x.get('param', '') for x in governance.get('blocked_changes', []) if isinstance(x, dict)]
        print(f"[governance] blocked locked overrides: {', '.join(blocked_params)}")
    print(f"[governance] code_fingerprint={governance.get('effective_code_fingerprint', '')}")

    records = load_force_records(args.force_file)
    segments = build_segments(config)

    mapper = ForceMapper(safety_level=args.safety_level)
    mapped = mapper.map_segments_symmetric(segments, records, mode=args.mode, top_k=args.top_k)

    solver = HybridNSGA2Solver()
    detail_solver = StructuralSolver()

    # 同类型仅设计一次，结果映射回所有同类型分块
    type_source = mapped['type_source']
    type_results: Dict[str, Any] = {}
    type_exhaustive: Dict[str, Any] = {}  # 穷举结果（按类型）
    type_multi_run_stats: Dict[str, Any] = {}  # 多次运行统计（按类型）
    for stype, seg_id in type_source.items():
        envelope = mapped['by_segment'].get(seg_id, [])
        if not envelope:
            type_results[stype] = None
            continue

        plot_path = os.path.join(os.path.dirname(args.report_md), f"pareto_{stype}.png")

        # Phase 2: 穷举枚举（在 NSGA-II 之前）
        exhaustive_result = None
        exhaustive_front_inds = None
        if args.exhaustive:
            exhaustive_result = solver.exhaustive_enumerate(section_params, envelope)
            exhaustive_front_inds = exhaustive_result.get('pareto_front', [])
            type_exhaustive[stype] = exhaustive_result

        # Phase 3: 多次运行 or 单次运行
        if args.n_runs > 1:
            multi_result = solver.solve_multi_run(
                section_params, envelope,
                n_runs=args.n_runs,
                seed=args.seed,
                plot_path=plot_path,
                exhaustive_front=exhaustive_front_inds,
            )
            type_multi_run_stats[stype] = multi_result.get('stats', {})
            best_run = multi_result.get('best_run')
            if not best_run:
                type_results[stype] = None
                continue
            main_result = best_run['result']
        else:
            main_result = solver.solve(
                section_params, envelope,
                plot_path=plot_path,
                seed=args.seed,
                exhaustive_front=exhaustive_front_inds,
            )

        if not main_result:
            type_results[stype] = None
            continue

        # 获取自适应目标标志
        use_sf_obj = main_result.get('use_sf_objective', False)

        max_main_candidates = int(section_params.get('joint_main_candidate_limit', 24))
        main_candidates = _build_main_scheme_candidates(main_result, max_count=max_main_candidates)
        if not main_candidates:
            type_results[stype] = None
            continue

        combo_candidates = []
        for main_scheme in main_candidates:
            detail = detail_solver.design_stirrups(main_scheme, section_params, envelope)
            detail['joint_objs'] = _joint_objectives(detail)
            combo_candidates.append(detail)

        joint_front = _joint_pareto(combo_candidates)

        # 对同一个 joint_front，用4种偏好分别评分选解
        PREFERENCES = ["cost", "safety", "construction", "balanced"]
        pref_results = {}
        # 收集偏好选中解（用于帕累托图高亮）
        pref_selections_for_plot = {}
        for pref in PREFERENCES:
            chosen_raw = _select_joint_best(joint_front, preference=pref)
            if not chosen_raw:
                continue
            joint_selection = _build_joint_selection_payload(joint_front, chosen_raw, preference=pref)
            chosen = dict(chosen_raw)
            chosen['joint_pareto'] = _build_joint_pareto_payload(joint_front, selected_raw=chosen_raw)
            chosen['joint_selection'] = joint_selection
            chosen['stirrup_selection'] = _build_stirrup_selection_payload(chosen)
            chosen.pop('joint_objs', None)
            chosen['use_sf_objective'] = use_sf_obj
            chosen['objective_mode'] = 'sf' if use_sf_obj else 'crack'
            pref_results[pref] = chosen

        if not pref_results:
            type_results[stype] = None
            continue

        # 附加穷举和统计信息到结果
        if exhaustive_result:
            for pref in pref_results:
                pref_results[pref]['exhaustive'] = {
                    'total': exhaustive_result.get('total', 0),
                    'valid_count': exhaustive_result.get('valid_count', 0),
                    'pareto_size': len(exhaustive_result.get('pareto_front', [])),
                }
        if stype in type_multi_run_stats:
            for pref in pref_results:
                pref_results[pref]['multi_run_stats'] = type_multi_run_stats[stype]

        type_results[stype] = pref_results

    write_report_md(
        args.report_md,
        config,
        args.config,
        args.force_file,
        args.mode,
        args.top_k,
        segments,
        type_source,
        type_results,
        governance=governance
    )
    write_report_csv(
        args.report_csv,
        segments,
        type_source,
        type_results,
        governance=governance
    )
    if args.report_xlsx:
        write_report_xlsx(
            args.report_xlsx,
            segments,
            type_source,
            type_results,
            governance=governance
        )

    paper_report_path = os.path.join(os.path.dirname(args.report_md), 'paper_report.md')
    write_paper_report_md(paper_report_path, config, segments, type_source, type_results)
    print(f'Paper Report: {paper_report_path}')

    PREF_NAMES = {"cost": "造价优先", "safety": "安全优先", "construction": "施工优先", "balanced": "均衡方案"}
    PREFERENCES = ["cost", "safety", "construction", "balanced"]
    print('\n=== Symmetric Design Summary (Multi-Preference) ===')
    for pref in PREFERENCES:
        print(f'\n--- {PREF_NAMES[pref]} ({pref.upper()}) ---')
        for seg in segments:
            seg_id = getattr(seg, 'segment_id', None)
            seg_type = getattr(seg, 'segment_type', None)
            if seg_id is None or seg_type is None:
                continue
            pref_data = type_results.get(seg_type)
            if not pref_data or not isinstance(pref_data, dict) or pref not in pref_data:
                print(f'{seg_id}: no solution')
                continue
            result = pref_data[pref]
            summary = result.get('summary', {})
            stirrup = _format_stirrup_spec(result.get('stirrup', {}))
            dist = _format_dist_spec(result.get('dist_rebar', {}))
            print(f'{seg_id}: {summary.get("type")} x {summary.get("count")}, stirrup={stirrup}, dist={dist}')

    print('\nType Representative Blocks:')
    for stype, seg_id in type_source.items():
        print(f'  {stype}: {seg_id}')
    print(f'\nReport: {args.report_md}')
    print(f'Report CSV: {args.report_csv}')
    if args.report_xlsx:
        print(f'Report XLSX: {args.report_xlsx}')


if __name__ == '__main__':
    main()
